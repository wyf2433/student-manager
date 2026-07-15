"""Excel 解析服务(学生名册 + 成绩导入)"""

import io

from openpyxl import load_workbook

from config import MAX_EXCEL_SIZE


def validate_excel_size(file_size: int) -> None:
    if file_size > MAX_EXCEL_SIZE:
        raise ValueError(f"文件过大,最大 {MAX_EXCEL_SIZE // 1024 // 1024}MB")


def parse_student_excel(file_bytes: bytes) -> list:
    """解析学生名册 Excel,返回 [{name, student_no, gender}]"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        raise ValueError("Excel 无有效数据")

    header = [str(cell).strip() if cell else "" for cell in rows[0]]
    name_idx = _find_column(header, ["姓名", "名字", "学生姓名"])
    if name_idx is None:
        wb.close()
        raise ValueError("未找到'姓名'列")

    student_no_idx = _find_column(header, ["学号", "考号", "考号/学号"])
    gender_idx = _find_column(header, ["性别", "男女"])

    students = []
    for row in rows[1:]:
        if not row or not row[name_idx]:
            continue
        name = str(row[name_idx]).strip()
        if not name:
            continue
        student_no = str(row[student_no_idx]).strip() if student_no_idx is not None and row[student_no_idx] else None
        gender = None
        if gender_idx is not None and row[gender_idx]:
            g = str(row[gender_idx]).strip()
            if g in ("男", "male", "M"):
                gender = "男"
            elif g in ("女", "female", "F"):
                gender = "女"
        students.append({"name": name, "student_no": student_no, "gender": gender})

    wb.close()
    if not students:
        raise ValueError("未解析到有效学生数据")
    return students


def parse_score_excel(file_bytes: bytes) -> dict:
    """解析成绩 Excel,返回 {exam_name, subjects, students}"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        raise ValueError("Excel 无有效数据")

    header = [str(cell).strip() if cell else "" for cell in rows[0]]
    name_idx = _find_column(header, ["姓名", "名字", "学生姓名"])
    if name_idx is None:
        wb.close()
        raise ValueError("未找到'姓名'列")

    exam_name_idx = _find_column(header, ["考试名称", "考试"])
    exam_name = None
    if exam_name_idx is not None and rows[1][exam_name_idx]:
        exam_name = str(rows[1][exam_name_idx]).strip()

    score_cols = {}
    for i, h in enumerate(header):
        if i == name_idx or i == exam_name_idx:
            continue
        if h and h not in ("学号", "考号", "性别", "班级", "年级", "排名", "总分", "名次", "班次", "校次", "班排", "校排", "级排"):
            score_cols[h] = i

    subjects = list(score_cols.keys())
    if not subjects:
        wb.close()
        raise ValueError("未找到科目列(物理/语文/数学等)")

    students = []
    for row in rows[1:]:
        if not row or not row[name_idx]:
            continue
        name = str(row[name_idx]).strip()
        if not name:
            continue
        scores = {}
        for subject, col_idx in score_cols.items():
            val = row[col_idx] if col_idx < len(row) and row[col_idx] is not None else None
            if val is not None:
                try:
                    scores[subject] = float(val)
                except (ValueError, TypeError):
                    scores[subject] = None
            else:
                scores[subject] = None
        students.append({"name": name, "scores": scores})

    wb.close()
    if not students:
        raise ValueError("未解析到有效成绩数据")

    return {
        "exam_name": exam_name,
        "subjects": subjects,
        "students": students,
        "total_students": len(students),
    }


def _find_column(header: list, candidates: list) -> int | None:
    for i, h in enumerate(header):
        for c in candidates:
            if c in h:
                return i
    return None
